import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from urllib.parse import urljoin
import urllib3
import time

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

BASE_URL = 'https://cpkrus.ru'
HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8',
    'Accept-Language': 'ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7',
}

def get_soup(url):
    response = requests.get(url, headers=HEADERS, verify=False, timeout=15)
    response.encoding = 'utf-8'
    return BeautifulSoup(response.text, 'html.parser')

def get_categories():
    """Сбор категорий (ссылка из кнопки 'Подробнее')"""
    soup = get_soup('https://cpkrus.ru/tyumen/kursy-professionalnoj-perepodgotovki/')
    categories = []
    tiles = soup.select('div.directions-module__item.direction-tile')
    print(f"Найдено блоков категорий: {len(tiles)}")
    for tile in tiles:
        detail_button = tile.select_one('a.custom-button--gray-o')
        if not detail_button or not detail_button.get('href'):
            continue
        name_tag = tile.select_one('div.direction-tile__name a')
        name = name_tag.get_text(strip=True) if name_tag else 'Без названия'
        link = detail_button.get('href')
        full_link = urljoin(BASE_URL, link)
        categories.append({'name': name, 'url': full_link})
        print(f"  Найдена категория: {name}")
    return categories

def extract_char_from_course_tile(course, label):
    """Извлечение характеристики из course-tile по тексту метки"""
    label_elem = course.find('div', class_='course-tile__characteristics-label', string=lambda t: t and label in t)
    if label_elem:
        value_elem = label_elem.find_next_sibling('div', class_='course-tile__characteristics-value')
        if value_elem:
            return value_elem.get_text(strip=True)
    return ''

def parse_courses_from_tiles(soup):
    """Парсинг программ из блоков .course-tile (обычный список)"""
    courses = []
    for course in soup.select('div.course-tile'):
        title_tag = course.select_one('div.course-tile__title a')
        if not title_tag:
            continue
        title = title_tag.get_text(strip=True)
        link = title_tag.get('href')
        full_link = urljoin(BASE_URL, link)

        volume = extract_char_from_course_tile(course, 'Объем программы')
        duration = extract_char_from_course_tile(course, 'Срок обучения')
        document = extract_char_from_course_tile(course, 'Получаемый документ')
        education = extract_char_from_course_tile(course, 'Требуемое образование')

        price_full = ''
        price_block = course.select_one('.course-tile__prices--body:first-child .course-tile__prices-current')
        if price_block:
            price_full = price_block.get_text(strip=True)

        courses.append({
            'title': title,
            'volume': volume,
            'duration': duration,
            'document': document,
            'education': education,
            'price_full': price_full,
            'url': full_link
        })
    return courses

def parse_single_program_page(soup, category_url):
    """Парсинг страницы, где только одна программа (структура courseinfo-wrap)"""
    courses = []
    # Название программы (обычно из h1, но можно уточнить)
    title_tag = soup.select_one('h1')
    title = title_tag.get_text(strip=True) if title_tag else 'Без названия'
    # Ссылка на программу (текущий URL)
    full_link = category_url

    # Объём программы (берём все вкладки)
    volume_tabs = soup.select('.program-volume__tab')
    volume = ', '.join([tab.get_text(strip=True) for tab in volume_tabs]) if volume_tabs else ''

    # Срок обучения (на страницах с одной программой он может отсутствовать, оставляем пустым)
    duration = ''

    # Получаемый документ (ищем активную вкладку "Диплом")
    document = ''
    doc_panel = soup.select_one('.documents-section__panel--active .documents-section__description')
    if doc_panel and 'диплом' in doc_panel.get_text().lower():
        document = 'Диплом о профессиональной переподготовке'
    else:
        document = 'Диплом'

    # Требуемое образование (ищем по тексту на странице)
    education = ''
    edu_text = soup.find(string=lambda t: t and 'Требуемое образование' in t)
    if edu_text:
        parent = edu_text.find_parent()
        if parent:
            value = parent.find_next_sibling()
            if value:
                education = value.get_text(strip=True)
    if not education:
        education = 'Высшее или среднее профессиональное'

    # Стоимость за весь курс (активная вкладка)
    price_full = ''
    active_price = soup.select_one('.course-price__content.active .course-price__price')
    if active_price:
        price_full = active_price.get_text(strip=True)

    courses.append({
        'title': title,
        'volume': volume,
        'duration': duration,
        'document': document,
        'education': education,
        'price_full': price_full,
        'url': full_link
    })
    return courses

def get_courses_from_category(category_url):
    """Определяет тип страницы и возвращает список программ"""
    soup = get_soup(category_url)
    # Сначала проверяем, есть ли блоки .course-tile (несколько программ)
    if soup.select('div.course-tile'):
        return parse_courses_from_tiles(soup)
    # Если нет — возможно, это страница с одной программой (courseinfo-wrap)
    elif soup.select_one('.courseinfo-wrap'):
        return parse_single_program_page(soup, category_url)
    else:
        print(f"  Не удалось найти программы на {category_url}")
        return []

def main():
    categories = get_categories()
    if not categories:
        print("Категории не найдены. Проверьте селекторы или доступ к сайту.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Обучающие программы"
    headers = ['Категория', 'Наименование программы', 'Объем программы',
               'Срок обучения', 'Получаемый документ', 'Требуемое образование',
               'Стоимость за весь курс', 'Ссылка на программу']
    ws.append(headers)

    row_num = 2
    for cat in categories:
        print(f'\nОбработка категории: {cat["name"]} ({cat["url"]})')
        courses = get_courses_from_category(cat['url'])
        print(f'  Найдено программ: {len(courses)}')
        for c in courses:
            ws.cell(row=row_num, column=1, value=cat['name'])
            ws.cell(row=row_num, column=2, value=c['title'])
            ws.cell(row=row_num, column=3, value=c['volume'])
            ws.cell(row=row_num, column=4, value=c['duration'])
            ws.cell(row=row_num, column=5, value=c['document'])
            ws.cell(row=row_num, column=6, value=c['education'])
            ws.cell(row=row_num, column=7, value=c['price_full'])
            ws.cell(row=row_num, column=8, value=c['url'])
            row_num += 1
        # Задержка между категориями
        time.sleep(2)

    # Автоширина колонок
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width

    wb.save('courses.xlsx')
    print(f'\nГотово. Сохранено {row_num - 2} записей в courses.xlsx')

if __name__ == '__main__':
    main()